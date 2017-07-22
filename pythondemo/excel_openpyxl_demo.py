# python 操作excel库 >>> openpyxl
# 官网教程：http://openpyxl.readthedocs.io/en/default/tutorial.html#
# 博客教程：http://openpyxl.readthedocs.io/en/default/tutorial.html#

import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
ws = wb.active
ws.title = 'openpyxl_test'

ws['A1'] = 42
ws.append([1, 2, 3])

ws['B1'] = datetime.datetime.now()

# wb.save('sample.xlsx')
# print('ok')

# reading and existing workbook
wb1 = load_workbook('sample.xlsx')
sheet_ranges = wb['openpyxl_test']
print(sheet_ranges['A1'].value)

# 总结
# 1. excel的创建 单元格的赋值 保存文件 从已有的excel读取内容